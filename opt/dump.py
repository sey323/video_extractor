import openpyxl

class Dumper:
    def __init__(self ,save_path,img_size = [160,90]):
        self.dump_text = ""
        self.bootstrap_cdn ='<head>\n \
                                <link rel="stylesheet" href="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.6/css/bootstrap.min.css" />\n \
                                <script src="https://ajax.googleapis.com/ajax/libs/jquery/1.11.1/jquery.min.js"></script>\n \
                                <script src="https://maxcdn.bootstrapcdn.com/bootstrap/3.3.6/js/bootstrap.min.js"></script>\n \
                            </head>'
        self.begin = '<table class="table table-responsive">'
        self.end = '</table>'
        self.content = ''
        self.img_size = img_size
        # openpyxl
        self.work_book = openpyxl.Workbook()
        self.work_sheet = self.work_book.worksheets[0]
        self.work_sheet.column_dimensions['C'].width = img_size[0] * 0.125
        self.work_sheet.column_dimensions['D'].width = 30

        # 初期設定
        self.work_sheet.cell(row=1,column=1).value = "No."
        self.work_sheet.cell(row=1,column=1).alignment = openpyxl.styles.Alignment(vertical = 'center',horizontal='center')
        self.work_sheet.cell(row=1,column=2).value =  "時間"
        self.work_sheet.cell(row=1,column=2).alignment = openpyxl.styles.Alignment(vertical = 'center',horizontal='center')
        self.work_sheet.cell(row=1,column=3).value =  "場面"
        self.work_sheet.cell(row=1,column=3).alignment = openpyxl.styles.Alignment(vertical = 'center',horizontal='center')
        self.work_sheet.cell(row=1,column=4).value = "撮影対象物"
        self.work_sheet.cell(row=1,column=4).alignment = openpyxl.styles.Alignment(vertical = 'center',horizontal='center')

        self.no = 1

        self.save_path = save_path

    def add_scene(self,frame,time,param_path ):
        '''
        フレームごとのHTML構文を追加する．
         Parameters
        ----------
        frame : int
            検出したフレーム
        '''

        # HTMLに出力
        with open(param_path, "r",encoding='utf-8') as f:
            param_data = f.read()
        self.content += '<tr class="col-md-12">\n \
                            <td class="col-md-1">{0}</td>\n \
                            <td class="col-md-1">{1}</td>\n \
                            <td class="col-md-5">\n  \
                                <img class="gen-img" src="{2}">\n \
                            </td>\n \
                            <td class="col-md-4">{3}</td>\n \
                            <td class="col-md-2"></td>\n \
                        </tr>'.format(self.no, time,"img/"+str(frame)+".jpg",param_data)

        # exelに出力
        exel_no = self.no+1
        self.work_sheet.cell(row=exel_no,column=1).value = self.no
        self.work_sheet.cell(row=exel_no,column=1).alignment = openpyxl.styles.Alignment(vertical = 'center')
        self.work_sheet.cell(row=exel_no,column=2).value = time
        self.work_sheet.cell(row=exel_no,column=2).alignment = openpyxl.styles.Alignment(vertical = 'center')
        self.work_sheet.cell(row=exel_no,column=4).value = param_data.replace('<br>', '\n')
        self.work_sheet.cell(row=exel_no,column=4).alignment = openpyxl.styles.Alignment(vertical = 'center', wrapText=True)

        # Exelに画像の入力
        self.work_sheet.row_dimensions[exel_no].height = self.img_size[1] * 0.78
        img = openpyxl.drawing.image.Image(self.save_path + "/exel/"+str(frame)+".jpg")
        img.anchor = self.work_sheet.cell(row=exel_no,column=3)
        img.anchor = 'C'+str(exel_no)
        self.work_sheet.add_image(img)

        self.no += 1

    def save_html( self  ):
        '''
        HTML形式で保存する．
        '''
        html = self.begin + self.content + self.end
        with open(self.save_path+'/index.html', mode='w') as f:
            f.write(html)

        self.work_book.save(self.save_path+'/result.xlsx')
        print('html save complete!')
